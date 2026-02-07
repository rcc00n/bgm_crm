from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Dict, Iterable, List, Optional, Tuple

from django.conf import settings
from django.utils.text import slugify

from .forms_store import parse_specs_text
from .models import Category, Product, ProductOption

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - handled at runtime for xlsx imports
    load_workbook = None

SKU_MAX_LEN = 64
SLUG_MAX_LEN = 200
PRICE_QUANT = Decimal("0.01")


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _normalize_row(row: Dict[str, object]) -> Dict[str, object]:
    normalized: Dict[str, object] = {}
    for key, value in row.items():
        if key is None:
            continue
        normalized_key = _normalize_header(key)
        if normalized_key and normalized_key not in normalized:
            normalized[normalized_key] = value
    return normalized


def _get_value(row: Dict[str, object], *aliases: str) -> Optional[object]:
    for alias in aliases:
        normalized = _normalize_header(alias)
        if normalized in row:
            value = row[normalized]
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            return value
    return None


def _parse_decimal(value: object) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.replace(",", "")
    raw = re.sub(r"[^\d\.\-]", "", raw)
    if not raw:
        return None
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def _clean_sku(value: object) -> str:
    if value is None:
        return ""
    sku = str(value).strip()
    if not sku:
        return ""
    return sku[:SKU_MAX_LEN]


def _sku_seed(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    seed = slugify(raw)
    return seed[:SKU_MAX_LEN]


def _generate_unique_sku(seed: object, taken: set[str]) -> str:
    base = _sku_seed(seed) or "product"
    base = base[:SKU_MAX_LEN]
    candidate = base
    counter = 2
    while candidate in taken:
        suffix = f"-{counter}"
        candidate = f"{base[:SKU_MAX_LEN - len(suffix)]}{suffix}"
        counter += 1
    taken.add(candidate)
    return candidate


def _slug_seed(value: object) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    seed = slugify(raw)
    return seed[:SLUG_MAX_LEN]


def _generate_unique_slug(seed: object, taken: set[str]) -> str:
    base = _slug_seed(seed) or "product"
    base = base[:SLUG_MAX_LEN]
    candidate = base
    counter = 2
    while candidate in taken:
        suffix = f"-{counter}"
        candidate = f"{base[:SLUG_MAX_LEN - len(suffix)]}{suffix}"
        counter += 1
    taken.add(candidate)
    return candidate


def _parse_int(value: object) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    raw = str(value).strip()
    if not raw:
        return None
    raw = re.sub(r"[^\d\-]", "", raw)
    if not raw:
        return None
    try:
        return int(raw)
    except ValueError:
        return None


def _parse_bool(value: object, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    raw = str(value).strip().lower()
    if raw in {"true", "1", "yes", "y", "published", "active", "enabled"}:
        return True
    if raw in {"false", "0", "no", "n", "draft", "inactive", "disabled"}:
        return False
    return default


def _parse_tags(value: object) -> List[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        tags = [str(item).strip() for item in value if str(item).strip()]
    else:
        raw = str(value)
        if "," in raw:
            tags = [item.strip() for item in raw.split(",")]
        elif ";" in raw:
            tags = [item.strip() for item in raw.split(";")]
        else:
            tags = [raw.strip()]
    deduped = []
    seen = set()
    for tag in tags:
        cleaned = tag[:32]
        if cleaned and cleaned not in seen:
            deduped.append(cleaned)
            seen.add(cleaned)
    return deduped


def _parse_specs(value: object) -> Dict[str, object]:
    if value is None:
        return {}
    if isinstance(value, dict):
        return value
    raw = str(value).strip()
    if not raw:
        return {}
    if raw.startswith("{") and raw.endswith("}"):
        try:
            return json.loads(raw)
        except Exception:
            return parse_specs_text(raw)
    return parse_specs_text(raw)


def _trim_text(value: Optional[object], max_len: int) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) <= max_len:
        return text
    return text[:max_len].rstrip()


def _read_csv(uploaded_file) -> Tuple[List[str], List[Dict[str, object]]]:
    raw = uploaded_file.read()
    uploaded_file.seek(0)
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("cp1252", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = reader.fieldnames or []
    rows = [row for row in reader]
    return headers, rows


def _read_xlsx(uploaded_file) -> Tuple[List[str], List[Dict[str, object]]]:
    if load_workbook is None:
        raise ValueError("openpyxl is required to import .xlsx files.")
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    rows: List[Dict[str, object]] = []
    for row in rows_iter:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        row_dict = {headers[idx]: row[idx] for idx in range(len(headers))}
        rows.append(row_dict)
    return headers, rows


def read_spreadsheet(uploaded_file) -> Tuple[List[str], List[Dict[str, object]]]:
    name = (uploaded_file.name or "").lower()
    if name.endswith(".csv"):
        return _read_csv(uploaded_file)
    if name.endswith(".xlsx"):
        return _read_xlsx(uploaded_file)
    raise ValueError("Unsupported file type. Upload a .csv or .xlsx file.")


def detect_shopify(headers: Iterable[str]) -> bool:
    normalized = {_normalize_header(header) for header in headers}
    required = {"handle", "title"}
    # Shopify exports always have Handle/Title, but variant columns differ by source.
    #
    # DDC "Shopify-like" files omit Variant SKU / Variant Price and instead use SKU
    # and RETAIL PRICE, plus Shopify-ish image/option columns. Auto mode should
    # still route these through the Shopify importer to avoid creating junk
    # products from image-only rows.
    markers = {
        # canonical Shopify
        "variantsku",
        "variantprice",
        "variantimage",
        # DDC/Shopify-like
        "sku",
        "imagesrc",
        "imageposition",
        "option1name",
        "option1value",
        # common Shopify-ish metadata
        "vendor",
        "bodyhtml",
        "published",
        "status",
        "tags",
        "type",
    }
    return required.issubset(normalized) and bool(markers & normalized)


def _ensure_unique_slug(base: str) -> str:
    max_len = 50
    base_slug = (base or "category")[:max_len]
    slug = base_slug
    counter = 2
    while Category.objects.filter(slug=slug).exists():
        suffix = f"-{counter}"
        trimmed = base_slug[: max_len - len(suffix)]
        slug = f"{trimmed}{suffix}"
        counter += 1
    return slug


@dataclass
class ImportResult:
    created_products: int = 0
    updated_products: int = 0
    skipped_products: int = 0
    created_options: int = 0
    updated_options: int = 0
    skipped_options: int = 0
    created_categories: int = 0
    errors: List[str] = field(default_factory=list)


FIELD_ALIASES = {
    "sku": [
        "sku",
        "variant sku",
        "product sku",
        "item",
        "item number",
        "part number",
        "partnumber",
        "mpn",
        "model",
        "product code",
        "code",
    ],
    "name": [
        "name",
        "title",
        "product",
        "product name",
        "item name",
    ],
    "description": [
        "description",
        "body",
        "body html",
        "body (html)",
        "details",
        "long description",
    ],
    "short_description": [
        "short description",
        "shortdescription",
        "excerpt",
        "summary",
        "seo description",
    ],
    "category": [
        "category",
        "product category",
        "type",
        "collection",
        "collections",
    ],
    "price": [
        "price",
        "unit price",
        "unitprice",
        "list price",
        "listprice",
        "msrp",
        "retail",
        "retail price",
        "variant price",
    ],
    "currency": [
        "currency",
        "currency code",
        "curr",
    ],
    "inventory": [
        "inventory",
        "inventory qty",
        "inventoryqty",
        "qty",
        "quantity",
        "stock",
        "on hand",
        "available",
    ],
    "tags": [
        "tags",
        "tag",
        "keywords",
    ],
    "specs": [
        "specs",
        "specifications",
        "attributes",
    ],
    "active": [
        "active",
        "is active",
        "published",
        "status",
    ],
    "vendor": [
        "vendor",
        "brand",
        "manufacturer",
    ],
}


REGION_PRICE_ALIASES = {
    "CAD": ["price / canada", "price canada", "price cad", "cad price"],
    "USD": [
        "price / united states",
        "price / us",
        "price / usa",
        "price usd",
        "usd price",
    ],
    "AUD": ["price / australia", "price australia", "price aud", "aud price"],
    "MXN": ["price / mexico", "price mexico", "price mxn", "mxn price"],
    "INT": ["price / international", "price international", "price intl"],
}


def _pick_price(row_norm: Dict[str, object], default_currency: str) -> Optional[object]:
    currency_key = (default_currency or "").upper()
    for alias in REGION_PRICE_ALIASES.get(currency_key, []):
        value = _get_value(row_norm, alias)
        if value is not None:
            return value
    return _get_value(row_norm, *FIELD_ALIASES["price"])


def _resolve_category(
    name: Optional[str],
    *,
    default_category: Optional[Category],
    create_missing: bool,
    dry_run: bool,
) -> Tuple[Optional[Category], bool]:
    if name:
        name = _trim_text(name, 120)
        existing = Category.objects.filter(name__iexact=name).first()
        if existing:
            return existing, False
        slug = _ensure_unique_slug(slugify(name))
        if create_missing:
            if dry_run:
                return None, True
            return Category.objects.create(name=name, slug=slug), True
    if default_category:
        return default_category, False
    return None, False


def import_products(
    *,
    uploaded_file,
    mode: str = "auto",
    default_category: Optional[Category] = None,
    default_currency: Optional[str] = None,
    update_existing: bool = False,
    create_missing_categories: bool = True,
    dieselr_foreign: bool = False,
    dry_run: bool = False,
    import_batch=None,
) -> ImportResult:
    headers, rows = read_spreadsheet(uploaded_file)
    if not headers or not rows:
        return ImportResult(errors=["The import file has no data rows."])

    file_name = (uploaded_file.name or "").lower()
    dieselr_from_name = "dieselr" in file_name
    price_multiplier = Decimal("1.4") if (dieselr_foreign or dieselr_from_name) else Decimal("1.0")

    auto_mode = mode == "auto"
    if auto_mode:
        mode = "shopify" if detect_shopify(headers) else "simple"

    if mode == "shopify":
        return _import_shopify(
            rows,
            default_category=default_category,
            default_currency=default_currency,
            price_multiplier=price_multiplier,
            update_existing=update_existing,
            create_missing_categories=create_missing_categories,
            dry_run=dry_run,
            import_batch=import_batch,
        )
    if mode == "simple":
        return _import_simple(
            rows,
            default_category=default_category,
            default_currency=default_currency,
            price_multiplier=price_multiplier,
            update_existing=update_existing,
            create_missing_categories=create_missing_categories,
            dry_run=dry_run,
            import_batch=import_batch,
        )
    return ImportResult(errors=[f"Unknown import mode: {mode}."])


def _build_sku_registry() -> set[str]:
    product_skus = set(Product.objects.values_list("sku", flat=True))
    option_skus = set(
        ProductOption.objects.exclude(sku__isnull=True)
        .exclude(sku="")
        .values_list("sku", flat=True)
    )
    return product_skus | option_skus


def _build_slug_registry() -> set[str]:
    return set(Product.objects.values_list("slug", flat=True))


def _apply_price_multiplier(price: Optional[Decimal], multiplier: Decimal) -> Optional[Decimal]:
    if price is None:
        return None
    return (price * multiplier).quantize(PRICE_QUANT)


def _import_simple(
    rows: Iterable[Dict[str, object]],
    *,
    default_category: Optional[Category],
    default_currency: Optional[str],
    price_multiplier: Decimal,
    update_existing: bool,
    create_missing_categories: bool,
    dry_run: bool,
    import_batch=None,
) -> ImportResult:
    result = ImportResult()
    currency_default = (default_currency or settings.DEFAULT_CURRENCY_CODE or "CAD").upper()
    taken_skus = _build_sku_registry()
    taken_slugs = _build_slug_registry()

    for idx, row in enumerate(rows, start=2):
        row_norm = _normalize_row(row)
        name_raw = _get_value(row_norm, *FIELD_ALIASES["name"])
        sku_raw = _get_value(row_norm, *FIELD_ALIASES["sku"])
        if sku_raw:
            sku = _clean_sku(sku_raw)
            if not sku:
                sku = _generate_unique_sku(name_raw or f"product-{idx}", taken_skus)
            else:
                taken_skus.add(sku)
        else:
            sku = _generate_unique_sku(name_raw or f"product-{idx}", taken_skus)

        name = _trim_text(name_raw, 180) if name_raw else _trim_text(sku, 180)
        if not name:
            result.errors.append(f"Row {idx}: missing product name.")
            result.skipped_products += 1
            continue

        category_raw = _get_value(row_norm, *FIELD_ALIASES["category"])
        category_name = str(category_raw).strip() if category_raw else ""
        category, category_created = _resolve_category(
            category_name,
            default_category=default_category,
            create_missing=create_missing_categories,
            dry_run=dry_run,
        )
        if category_created:
            result.created_categories += 1
        if not category and not dry_run:
            result.errors.append(f"Row {idx}: missing category and no default set.")
            result.skipped_products += 1
            continue

        price_raw = _pick_price(row_norm, currency_default)
        price = _parse_decimal(price_raw) if price_raw is not None else None
        price = _apply_price_multiplier(price, price_multiplier)
        currency_raw = _get_value(row_norm, *FIELD_ALIASES["currency"])
        currency = (str(currency_raw).strip().upper() if currency_raw else currency_default)
        inventory_raw = _get_value(row_norm, *FIELD_ALIASES["inventory"])
        inventory = _parse_int(inventory_raw) if inventory_raw is not None else None
        description_raw = _get_value(row_norm, *FIELD_ALIASES["description"])
        description = str(description_raw).strip() if description_raw else ""
        short_desc_raw = _get_value(row_norm, *FIELD_ALIASES["short_description"])
        short_description = _trim_text(short_desc_raw, 240) if short_desc_raw else ""
        tags_raw = _get_value(row_norm, *FIELD_ALIASES["tags"])
        tags = _parse_tags(tags_raw)
        specs_raw = _get_value(row_norm, *FIELD_ALIASES["specs"])
        specs = _parse_specs(specs_raw)
        active_raw = _get_value(row_norm, *FIELD_ALIASES["active"])
        is_active = _parse_bool(active_raw, default=True)
        vendor_raw = _get_value(row_norm, *FIELD_ALIASES["vendor"])
        if vendor_raw:
            vendor_tag = _trim_text(vendor_raw, 32)
            if vendor_tag and vendor_tag not in tags:
                tags.append(vendor_tag)

        existing = Product.objects.filter(sku=sku).first()
        if existing:
            if not update_existing:
                result.skipped_products += 1
                continue
            if dry_run:
                result.updated_products += 1
                continue
            try:
                if name_raw:
                    existing.name = name
                if category_raw or default_category:
                    if category:
                        existing.category = category
                if price is not None:
                    existing.price = price
                if currency_raw:
                    existing.currency = currency
                if inventory is not None:
                    existing.inventory = inventory
                if description_raw:
                    existing.description = description
                if short_description:
                    existing.short_description = short_description
                if tags_raw is not None:
                    existing.tags = tags
                if specs_raw is not None and specs:
                    existing.specs = specs
                if active_raw is not None:
                    existing.is_active = is_active
                existing.save()
            except Exception as exc:
                result.errors.append(f"Row {idx}: {exc}")
                result.skipped_products += 1
                continue
            else:
                result.updated_products += 1
            continue

        if dry_run:
            result.created_products += 1
            continue

        price_create = price if price is not None else Decimal("0.00")
        inventory_create = inventory if inventory is not None else 0
        slug_seed = name or sku
        slug_value = _generate_unique_slug(slug_seed, taken_slugs)

        try:
            Product.objects.create(
                name=name,
                sku=sku,
                slug=slug_value,
                category=category,
                price=price_create,
                currency=currency,
                inventory=inventory_create,
                is_active=is_active,
                short_description=short_description,
                description=description,
                tags=tags,
                specs=specs,
                import_batch=import_batch,
            )
        except Exception as exc:
            result.errors.append(f"Row {idx}: {exc}")
            result.skipped_products += 1
        else:
            result.created_products += 1

    return result


def _shopify_option_label(row_norm: Dict[str, object]) -> Optional[str]:
    parts = []
    for idx in (1, 2, 3):
        opt_name = _get_value(row_norm, f"option{idx} name")
        opt_value = _get_value(row_norm, f"option{idx} value")
        if not opt_value:
            continue
        value = str(opt_value).strip()
        if not value or value.lower() == "default title":
            continue
        if opt_name:
            name = str(opt_name).strip()
            if name and name.lower() not in {"title", "default"}:
                parts.append(f"{name}: {value}")
                continue
        parts.append(value)
    if not parts:
        return None
    return " / ".join(parts)[:120]


def _pick_shopify_primary_row(
    group: List[Dict[str, object]],
    *,
    currency_default: str,
) -> Dict[str, object]:
    """
    Shopify-like imports can include many rows per handle, including image-only rows
    with blank metadata. Pick the best row for product-level fields deterministically.
    """
    best = group[0]
    best_score = -1
    for row_norm in group:
        score = 0
        if _get_value(row_norm, "title") is not None:
            score += 10
        if _get_value(row_norm, "body (html)", "body html", "body") is not None:
            score += 6
        if _parse_decimal(_pick_price(row_norm, currency_default)) is not None:
            score += 3
        if _get_value(row_norm, "vendor") is not None:
            score += 2
        if _get_value(row_norm, "product category", "type", "collection") is not None:
            score += 2

        # Tie-breaker: earliest in file wins (keep existing best when scores match).
        if score > best_score:
            best = row_norm
            best_score = score
    return best


def _pick_shopify_main_image_url(group: List[Dict[str, object]]) -> str:
    """
    Extract the primary product image for a handle group.

    - Collect Image Src / Variant Image across all rows
    - Sort by Image Position (lowest wins); rows without a position sort last
    - Deduplicate URLs (keeping the best-positioned occurrence)
    """
    best_by_url: Dict[str, Tuple[int, int]] = {}
    for idx, row_norm in enumerate(group):
        src_raw = _get_value(row_norm, "image src", "variant image")
        if src_raw is None:
            continue
        src = str(src_raw).strip()
        if not src:
            continue
        pos = _parse_int(_get_value(row_norm, "image position"))
        sort_pos = pos if pos is not None else 10**9
        key = (sort_pos, idx)
        if src not in best_by_url or key < best_by_url[src]:
            best_by_url[src] = key
    if not best_by_url:
        return ""
    best_src, _best_key = min(best_by_url.items(), key=lambda kv: kv[1])
    return best_src


def _import_shopify(
    rows: Iterable[Dict[str, object]],
    *,
    default_category: Optional[Category],
    default_currency: Optional[str],
    price_multiplier: Decimal,
    update_existing: bool,
    create_missing_categories: bool,
    dry_run: bool,
    import_batch=None,
) -> ImportResult:
    result = ImportResult()
    currency_default = (default_currency or settings.DEFAULT_CURRENCY_CODE or "CAD").upper()
    grouped: Dict[str, List[Dict[str, object]]] = {}
    taken_skus = _build_sku_registry()
    taken_slugs = _build_slug_registry()

    for row in rows:
        row_norm = _normalize_row(row)
        handle_raw = _get_value(row_norm, "handle")
        title_raw = _get_value(row_norm, "title")
        handle = str(handle_raw).strip() if handle_raw else str(title_raw or "").strip()
        if not handle:
            result.errors.append("Shopify row missing handle/title.")
            continue
        grouped.setdefault(handle, []).append(row_norm)

    for handle, group in grouped.items():
        handle_slug = _slug_seed(handle)
        primary = _pick_shopify_primary_row(group, currency_default=currency_default)

        title_raw = _get_value(primary, "title")
        name = _trim_text(title_raw or handle, 180)
        if not name:
            result.errors.append(f"{handle}: missing title.")
            result.skipped_products += 1
            continue

        category_raw = _get_value(primary, "product category", "type", "collection")
        category_name = str(category_raw).strip() if category_raw else ""

        description_raw = _get_value(primary, "body (html)", "body html", "body")
        description = str(description_raw).strip() if description_raw else ""
        short_desc_raw = _get_value(primary, "seo description")
        short_description = _trim_text(short_desc_raw, 240)
        tags_raw = _get_value(primary, "tags")
        tags = _parse_tags(tags_raw)
        vendor_raw = _get_value(primary, "vendor")
        if vendor_raw:
            vendor_tag = _trim_text(vendor_raw, 32)
            if vendor_tag and vendor_tag not in tags:
                tags.append(vendor_tag)
        active_raw = _get_value(primary, "published", "status")
        is_active = _parse_bool(active_raw, default=True)

        variant_prices = []
        for row_norm in group:
            price_value = _pick_price(row_norm, currency_default)
            price = _parse_decimal(price_value)
            if price is not None:
                price = _apply_price_multiplier(price, price_multiplier)
                variant_prices.append(price)
        has_price = bool(variant_prices)
        base_price = min(variant_prices) if variant_prices else Decimal("0.00")

        base_sku = None
        for row_norm in group:
            opt_label = _shopify_option_label(row_norm)
            if not opt_label:
                sku_candidate = _get_value(row_norm, "variant sku", "sku")
                if sku_candidate:
                    base_sku = _clean_sku(sku_candidate)
                    break
        if not base_sku:
            base_sku = _generate_unique_sku(name or handle, taken_skus)
        else:
            taken_skus.add(base_sku)

        inventory_raw = _get_value(primary, "variant inventory qty", "inventory", "qty", "quantity")
        inventory_value = _parse_int(inventory_raw) if inventory_raw is not None else None

        main_image_url = _pick_shopify_main_image_url(group)

        # Prefer stable Shopify handle/slug matching; fall back to SKU.
        existing = Product.objects.filter(slug=handle_slug).first() if handle_slug else None
        if not existing:
            existing = Product.objects.filter(sku=base_sku).first()

        if existing:
            if not update_existing:
                result.skipped_products += 1
                continue
            if dry_run:
                result.updated_products += 1
                product = existing
            else:
                category = None
                category_created = False
                if category_raw is not None or default_category is not None:
                    category, category_created = _resolve_category(
                        category_name,
                        default_category=default_category,
                        create_missing=create_missing_categories,
                        dry_run=dry_run,
                    )
                if category_created:
                    result.created_categories += 1
                try:
                    if title_raw is not None:
                        existing.name = name
                    if (category_raw is not None or default_category is not None) and category:
                        existing.category = category
                    if has_price:
                        existing.price = base_price
                        existing.currency = currency_default
                    if inventory_value is not None:
                        existing.inventory = inventory_value
                    if description_raw is not None:
                        existing.description = description
                    if short_description:
                        existing.short_description = short_description
                    if tags_raw is not None:
                        existing.tags = tags
                    if active_raw is not None:
                        existing.is_active = is_active
                    if main_image_url:
                        existing.main_image = main_image_url
                    existing.save()
                except Exception as exc:
                    result.errors.append(f"{handle}: {exc}")
                    result.skipped_products += 1
                    continue
                else:
                    result.updated_products += 1
            product = existing
        else:
            category, category_created = _resolve_category(
                category_name,
                default_category=default_category,
                create_missing=create_missing_categories,
                dry_run=dry_run,
            )
            if category_created:
                result.created_categories += 1
            if not category and not dry_run:
                result.errors.append(f"{handle}: missing category and no default set.")
                result.skipped_products += 1
                continue

            if dry_run:
                result.created_products += 1
                product = None
            else:
                slug_seed = handle or name or base_sku
                slug_value = _generate_unique_slug(slug_seed, taken_slugs)
                try:
                    product = Product.objects.create(
                        name=name,
                        sku=base_sku,
                        slug=slug_value,
                        category=category,
                        price=base_price,
                        currency=currency_default,
                        inventory=inventory_value or 0,
                        is_active=is_active,
                        short_description=short_description,
                        description=description,
                        tags=tags,
                        main_image=main_image_url or None,
                        import_batch=import_batch,
                    )
                except Exception as exc:
                    result.errors.append(f"{handle}: {exc}")
                    result.skipped_products += 1
                    continue
                else:
                    result.created_products += 1

        for row_norm in group:
            opt_label = _shopify_option_label(row_norm)
            if not opt_label:
                continue
            opt_sku_raw = _get_value(row_norm, "variant sku", "sku")
            opt_sku = _clean_sku(opt_sku_raw) if opt_sku_raw else ""
            opt_price_raw = _pick_price(row_norm, currency_default)
            opt_price = _parse_decimal(opt_price_raw)
            opt_price = _apply_price_multiplier(opt_price, price_multiplier)
            opt_active_raw = _get_value(row_norm, "published", "status")
            opt_active = _parse_bool(opt_active_raw, default=True)

            existing_opt = None
            if opt_sku:
                existing_opt = ProductOption.objects.filter(sku=opt_sku).first()
            if not existing_opt and product:
                existing_opt = ProductOption.objects.filter(product=product, name=opt_label).first()

            if existing_opt:
                if not update_existing:
                    result.skipped_options += 1
                    continue
                if dry_run:
                    result.updated_options += 1
                    continue
                try:
                    existing_opt.name = opt_label
                    if opt_sku:
                        existing_opt.sku = opt_sku
                    if opt_price is not None:
                        existing_opt.price = opt_price
                    if opt_active_raw is not None:
                        existing_opt.is_active = opt_active
                    existing_opt.save()
                except Exception as exc:
                    result.errors.append(f"{handle}: option {opt_label} -> {exc}")
                    result.skipped_options += 1
                else:
                    result.updated_options += 1
                continue

            if dry_run:
                result.created_options += 1
                continue
            if not product:
                continue
            try:
                ProductOption.objects.create(
                    product=product,
                    name=opt_label,
                    sku=opt_sku or None,
                    price=opt_price,
                    is_active=opt_active,
                    import_batch=import_batch,
                )
            except Exception as exc:
                result.errors.append(f"{handle}: option {opt_label} -> {exc}")
                result.skipped_options += 1
            else:
                result.created_options += 1

    return result
