from __future__ import annotations

from dataclasses import dataclass
import csv
import io
import re
from email.utils import parseaddr

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db.models import Q
from django.utils import timezone

from openpyxl import load_workbook

from core.email_templates import base_email_context, join_text_sections
from core.emails import build_email_html, send_html_email
from core.models import (
    CustomUserDisplay,
    EmailCampaign,
    EmailCampaignRecipient,
    EmailSubscriber,
)


EMAIL_SPLIT_RE = re.compile(r"[\s,;]+")


class _SafeDict(dict):
    def __missing__(self, key: str) -> str:
        return ""


def _normalize_email(value: str | None) -> str:
    if not value:
        return ""
    return value.strip().lower()


def _format_value(value: str, context: dict[str, str]) -> str:
    if not value:
        return ""
    try:
        return value.format_map(_SafeDict(context))
    except Exception:
        return value


def _split_lines(value: str) -> list[str]:
    if not value:
        return []
    return [line.strip() for line in value.splitlines() if line.strip()]


def _render_lines(value: str, context: dict[str, str]) -> list[str]:
    return [_format_value(line, context) for line in _split_lines(value)]


@dataclass(frozen=True)
class CampaignContent:
    subject: str
    preheader: str
    text_body: str
    html_body: str


def render_campaign_email(campaign: EmailCampaign, extra_context: dict[str, object] | None = None) -> CampaignContent:
    context = base_email_context(extra_context)

    subject = _format_value(campaign.subject, context)
    preheader = _format_value(campaign.preheader, context)
    title = _format_value(campaign.title or campaign.name or campaign.subject, context)
    greeting = _format_value(campaign.greeting, context)
    intro_lines = _render_lines(campaign.intro, context)
    notice_title = _format_value(campaign.notice_title, context)
    notice_lines = _render_lines(campaign.notice, context)
    footer_lines = _render_lines(campaign.footer, context)
    cta_label = _format_value(campaign.cta_label, context)
    cta_url = _format_value(campaign.cta_url, context)

    html_body = build_email_html(
        title=title,
        preheader=preheader,
        greeting=greeting,
        intro_lines=intro_lines,
        notice_title=notice_title,
        notice_lines=notice_lines,
        footer_lines=footer_lines,
        cta_label=cta_label,
        cta_url=cta_url,
    )

    notice_section: list[str] = []
    if notice_title:
        notice_section.append(notice_title)
    notice_section.extend(notice_lines)

    cta_section: list[str] = []
    if cta_label or cta_url:
        target = cta_url or context.get("company_website", "")
        if target:
            cta_section.append(f"{cta_label + ': ' if cta_label else ''}{target}")
        elif cta_label:
            cta_section.append(cta_label)

    text_body = join_text_sections(
        greeting,
        intro_lines,
        notice_section,
        footer_lines,
        cta_section,
    )

    return CampaignContent(
        subject=subject,
        preheader=preheader,
        text_body=text_body,
        html_body=html_body,
    )


def estimate_campaign_audience(campaign: EmailCampaign) -> dict[str, int]:
    subscriber_count = 0
    user_count = 0
    if campaign.include_subscribers:
        subscriber_count = EmailSubscriber.objects.filter(is_active=True).count()
    if campaign.include_registered_users:
        user_count = (
            CustomUserDisplay.objects.filter(
                Q(userprofile__email_marketing_consent=True)
                | Q(userprofile__email_product_updates=True)
                | Q(userprofile__email_service_updates=True)
            )
            .exclude(email__isnull=True)
            .exclude(email__exact="")
            .count()
        )
    return {
        "subscriber_count": subscriber_count,
        "user_count": user_count,
        "estimated_total": subscriber_count + user_count,
    }


def collect_campaign_recipients(campaign: EmailCampaign) -> list[dict[str, object]]:
    recipients: dict[str, dict[str, object]] = {}

    if campaign.include_subscribers:
        for subscriber in EmailSubscriber.objects.filter(is_active=True):
            email = _normalize_email(subscriber.email)
            if not email:
                continue
            recipients[email] = {
                "email": email,
                "user": None,
                "source": EmailCampaignRecipient.Source.SUBSCRIBER,
            }

    if campaign.include_registered_users:
        users = (
            CustomUserDisplay.objects.filter(
                Q(userprofile__email_marketing_consent=True)
                | Q(userprofile__email_product_updates=True)
                | Q(userprofile__email_service_updates=True)
            )
            .exclude(email__isnull=True)
            .exclude(email__exact="")
            .select_related("userprofile")
        )
        for user in users:
            email = _normalize_email(user.email)
            if not email:
                continue
            if email in recipients:
                if recipients[email].get("user") is None:
                    recipients[email]["user"] = user
                recipients[email]["source"] = EmailCampaignRecipient.Source.USER
            else:
                recipients[email] = {
                    "email": email,
                    "user": user,
                    "source": EmailCampaignRecipient.Source.USER,
                }

    return list(recipients.values())


def send_campaign(
    campaign: EmailCampaign,
    *,
    triggered_by=None,
    force: bool = False,
) -> dict[str, int | str]:
    if not force and campaign.status not in {
        EmailCampaign.Status.DRAFT,
        EmailCampaign.Status.PARTIAL,
        EmailCampaign.Status.FAILED,
    }:
        return {"status": "skipped", "total": 0, "sent": 0, "failed": 0, "skipped_count": 0}

    recipients = collect_campaign_recipients(campaign)
    total = len(recipients)
    if total == 0:
        return {"status": "no_recipients", "total": 0, "sent": 0, "failed": 0, "skipped_count": 0}

    sender = (
        campaign.from_email
        or getattr(settings, "DEFAULT_FROM_EMAIL", "")
        or getattr(settings, "SUPPORT_EMAIL", "")
    )
    if not sender:
        raise ValueError("Missing DEFAULT_FROM_EMAIL/SUPPORT_EMAIL for campaign sending.")

    now = timezone.now()
    campaign.status = EmailCampaign.Status.SENDING
    campaign.send_started_at = campaign.send_started_at or now
    campaign.sent_by = triggered_by or campaign.sent_by
    campaign.recipients_total = total
    campaign.save(
        update_fields=["status", "send_started_at", "sent_by", "recipients_total", "updated_at"]
    )

    sent_count = 0
    failed_count = 0
    skipped_count = 0

    existing = {
        recipient.email: recipient
        for recipient in EmailCampaignRecipient.objects.filter(campaign=campaign)
    }

    for entry in recipients:
        email = str(entry.get("email") or "").strip().lower()
        if not email:
            continue

        record = existing.get(email)
        if record and record.status == EmailCampaignRecipient.Status.SENT and not force:
            skipped_count += 1
            continue

        user = entry.get("user")
        extra_context = {
            "email": email,
            "first_name": getattr(user, "first_name", "") if user else "",
            "last_name": getattr(user, "last_name", "") if user else "",
            "full_name": user.get_full_name() if user else "",
        }
        content = render_campaign_email(campaign, extra_context=extra_context)

        try:
            send_html_email(
                subject=content.subject,
                text_body=content.text_body,
                html_body=content.html_body,
                from_email=sender,
                recipient_list=[email],
                email_type=f"campaign:{campaign.pk}",
            )
        except Exception as exc:
            failed_count += 1
            error_message = str(exc)[:255]
            if record:
                record.status = EmailCampaignRecipient.Status.FAILED
                record.error_message = error_message
                record.sent_at = None
                record.user = user if user else record.user
                record.source = entry.get("source") or record.source
                record.save(
                    update_fields=["status", "error_message", "sent_at", "user", "source"]
                )
            else:
                EmailCampaignRecipient.objects.create(
                    campaign=campaign,
                    email=email,
                    user=user,
                    source=entry.get("source") or EmailCampaignRecipient.Source.SUBSCRIBER,
                    status=EmailCampaignRecipient.Status.FAILED,
                    error_message=error_message,
                )
            continue

        sent_count += 1
        if record:
            record.status = EmailCampaignRecipient.Status.SENT
            record.error_message = ""
            record.sent_at = timezone.now()
            record.user = user if user else record.user
            record.source = entry.get("source") or record.source
            record.save(
                update_fields=["status", "error_message", "sent_at", "user", "source"]
            )
        else:
            EmailCampaignRecipient.objects.create(
                campaign=campaign,
                email=email,
                user=user,
                source=entry.get("source") or EmailCampaignRecipient.Source.SUBSCRIBER,
                status=EmailCampaignRecipient.Status.SENT,
                sent_at=timezone.now(),
            )

    campaign.sent_count = sent_count
    campaign.failed_count = failed_count
    campaign.send_completed_at = timezone.now()
    if failed_count and sent_count:
        campaign.status = EmailCampaign.Status.PARTIAL
    elif failed_count and not sent_count:
        campaign.status = EmailCampaign.Status.FAILED
    else:
        campaign.status = EmailCampaign.Status.SENT
    campaign.save(
        update_fields=[
            "status",
            "sent_count",
            "failed_count",
            "send_completed_at",
            "updated_at",
        ]
    )

    return {
        "status": campaign.status,
        "total": total,
        "sent": sent_count,
        "failed": failed_count,
        "skipped_count": skipped_count,
    }


def _extract_emails_from_text(text: str) -> tuple[set[str], int]:
    emails: set[str] = set()
    invalid_count = 0

    if not text:
        return emails, invalid_count

    raw_candidates: list[str] = []
    parsed = parseaddr(text)
    if parsed and parsed[1]:
        raw_candidates.append(parsed[1])
    for part in EMAIL_SPLIT_RE.split(text):
        if "@" in part:
            raw_candidates.append(part.strip("<>\"'()[]{}"))

    for candidate in raw_candidates:
        candidate = candidate.strip()
        if not candidate or "@" not in candidate:
            continue
        normalized = _normalize_email(candidate)
        try:
            validate_email(normalized)
        except ValidationError:
            invalid_count += 1
            continue
        emails.add(normalized)

    return emails, invalid_count


def _extract_emails_from_csv(file_obj) -> tuple[set[str], int]:
    emails: set[str] = set()
    invalid_total = 0
    text = file_obj.read().decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    for row in reader:
        for cell in row:
            if cell is None:
                continue
            found, invalid = _extract_emails_from_text(str(cell))
            emails.update(found)
            invalid_total += invalid
    return emails, invalid_total


def _extract_emails_from_xlsx(file_obj) -> tuple[set[str], int]:
    emails: set[str] = set()
    invalid_total = 0
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if cell is None:
                continue
            found, invalid = _extract_emails_from_text(str(cell))
            emails.update(found)
            invalid_total += invalid
    return emails, invalid_total


def import_email_subscribers(
    file_obj,
    *,
    added_by=None,
    reactivate: bool = True,
) -> dict[str, int]:
    filename = (getattr(file_obj, "name", "") or "").lower()

    if filename.endswith(".xlsx") or filename.endswith(".xlsm") or filename.endswith(".xltx"):
        emails, invalid_count = _extract_emails_from_xlsx(file_obj)
    else:
        emails, invalid_count = _extract_emails_from_csv(file_obj)

    if not emails:
        return {
            "total": 0,
            "created": 0,
            "reactivated": 0,
            "skipped": 0,
            "invalid": invalid_count,
        }

    existing = EmailSubscriber.objects.filter(email__in=emails)
    existing_map = {subscriber.email: subscriber for subscriber in existing}

    new_subscribers = []
    for email in emails:
        if email in existing_map:
            continue
        new_subscribers.append(
            EmailSubscriber(
                email=email,
                source=EmailSubscriber.Source.IMPORT,
                added_by=added_by,
            )
        )

    created_count = 0
    if new_subscribers:
        EmailSubscriber.objects.bulk_create(new_subscribers)
        created_count = len(new_subscribers)

    reactivated_count = 0
    if reactivate and existing_map:
        reactivated_count = EmailSubscriber.objects.filter(
            email__in=emails, is_active=False
        ).update(is_active=True, updated_at=timezone.now())

    skipped_count = len(emails) - created_count - reactivated_count

    return {
        "total": len(emails),
        "created": created_count,
        "reactivated": reactivated_count,
        "skipped": skipped_count,
        "invalid": invalid_count,
    }
